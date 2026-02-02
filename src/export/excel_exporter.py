from __future__ import annotations

from pathlib import Path
from typing import Any
from datetime import datetime, time, timedelta
import re
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

from src.util.paths import resource_path
from copy import copy

import calendar
from datetime import date
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import Counter
from openpyxl import Workbook

from src.domain.time_rules import classify_dt_odt

TR_DOW = ["Pt", "Sa", "Çr", "Pş", "Cu", "Ct", "Pa"]  # Monday=0


def _pick_worksheet(wb: Workbook, preferred: str, *, contains_any: list[str] | None = None):
    """Template sayfa adı değişse bile çalışmak için esnek seçim.

    - preferred mevcutsa onu döner.
    - değilse contains_any (case-insensitive) anahtar kelimelerinden birini içeren ilk sayfayı döner.
    - en sonda ilk sayfayı döner.
    """
    if preferred in wb.sheetnames:
        return wb[preferred]
    if contains_any:
        keys = [k.lower() for k in contains_any if k]
        for n in wb.sheetnames:
            nl = (n or "").lower()
            if any(k in nl for k in keys):
                return wb[n]
    # fallback: ilk sayfa
    return wb[wb.sheetnames[0]]


def _row_idx_to_time(row_idx: int) -> time:
    """Grid satırı -> kuşak başlangıç saati.
    Şablon: 07:00-20:00, 15dk.
    """
    mins = 7 * 60 + int(row_idx) * 15
    return time(mins // 60, mins % 60)


def _norm_hour_label(label: str) -> str:
    """Saat etiketini tek formata indirger.

    Repository tarafındaki normalize mantığı ile aynı olmalı; aksi halde
    '08:00 - 09:00' gibi başlıklarda eşleşme kaçıyor.
    """
    s = (label or "").strip()
    s = s.replace("–", "-").replace("—", "-")
    # sonda '(...)' varsa at
    import re

    s = re.sub(r"\([^\)]*\)\s*$", "", s).strip()
    s = re.sub(r"\s*-\s*", "-", s)
    m = re.match(r"^(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})$", s)
    if m:
        h1, m1, h2, m2 = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        return f"{h1:02d}:{m1:02d}-{h2:02d}:{m2:02d}"
    return s

def export_excel(template_path: Path, out_path: Path, payload: dict[str, Any]) -> Path:
    if not template_path.exists():
        raise FileNotFoundError(f"Template bulunamadı: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    # Excel açılınca formüller yeniden hesaplasın
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    ws = _pick_worksheet(
        wb,
        "REZERVASYON ve PLANLAMA",
        contains_any=["rezervasyon", "planlama"],
    )

    # --- Header labels + values ---
    # Şablonda etiketler A sütununda, değerler B sütununda.
    ws["A1"].value = "Ajans:"
    ws["B1"].value = str(payload.get("agency_name", "")).strip()

    ws["A2"].value = "Reklam Veren:"
    ws["B2"].value = str(payload.get("advertiser_name", "")).strip()

    ws["A3"].value = "Ürün:"
    ws["B3"].value = str(payload.get("product_name", "")).strip()

    ws["A4"].value = "Plan Başlığı:"
    ws["B4"].value = str(payload.get("plan_title", "")).strip()

    ws["A5"].value = "Rezervasyon No:"
    ws["B5"].value = str(payload.get("reservation_no", "")).strip()

    # eski şablonlardan kalma değerler iki kere görünmesin
    for addr in ("C1", "C2", "C3", "C4", "C5"):
        try:
            ws[addr].value = None
        except Exception:
            pass
    
    # --- Header label font fix (A4/A5/A6 da A1 gibi kalın olsun) ---
    hdr_font = copy(ws["A1"].font)
    for addr in ("A4", "A5", "A6"):
        ws[addr].font = hdr_font


    # Dönemi: AY/YIL (plan_date varsa oradan üret)
    period = str(payload.get("period", "")).strip()
    if not period:
        plan_date = str(payload.get("plan_date", "")).strip()  # "YYYY-MM-DD" bekliyoruz
        try:
            y, m, _ = plan_date.split("-")
            period = f"{m}/{y}"
        except Exception:
            period = ""

    ws["A6"].value = "Dönemi:"
    ws["B6"].value = period
    # eski şablonlarda C6/D6 dolu olabilir; iki kere yazmasın
    for addr in ("C6", "D6"):
        try:
            ws[addr].value = None
        except Exception:
            pass

    # --- Ay/Gün başlıklarını plan_date'e göre düzelt + hafta sonu sütunlarını boyala ---
    plan_date = str(payload.get("plan_date", "")).strip()  # "YYYY-MM-DD"
    year = month = None
    try:
        y, m, _ = plan_date.split("-")
        year = int(y)
        month = int(m)
    except Exception:
        year = None
        month = None

    if year and month:
        days_in_month = calendar.monthrange(year, month)[1]
        HEADER_ROW = 7          # C7..AG7
        FIRST_DAY_COL = 4       # D
        MAX_DAYS = 31

        GRID_START_ROW = 8
        GRID_ROWS = 52

        # --- Şablondan "weekday/weekend" fill örneklerini otomatik yakala ---
        # Not: DT satırlarının şablonda ayrı bir tonu var. Önceki yaklaşım tek bir satırdan
        # örnek aldığı için DT satırı tonunu ezip geçiyordu. Bu yüzden DT ve ODT için ayrı ayrı
        # örnekleyip, export sırasında satır tipine göre fill basıyoruz.

        # Grid alanında en sık görülen fill = weekday, daha az görülen = weekend
        def fill_sig_from_cell(cell):
            f = cell.fill
            # StyleProxy'yi hash'e sokmayacağız; sadece primitive imza
            pt = getattr(f, "patternType", None)
            fg = getattr(getattr(f, "fgColor", None), "rgb", None)
            bg = getattr(getattr(f, "bgColor", None), "rgb", None)
            return (pt, fg, bg)

        def analyze_grid_row(sample_row: int) -> tuple[Any, Any]:
            """Verilen satır için (weekday_fill, weekend_fill) döndürür."""
            cells = []
            for d in range(1, MAX_DAYS + 1):
                c = FIRST_DAY_COL + (d - 1)
                cells.append(ws.cell(sample_row, c))

            sigs = [fill_sig_from_cell(cell) for cell in cells]
            cnt = Counter(sigs).most_common()
            weekday_sig = cnt[0][0] if cnt else None
            weekend_sig = cnt[1][0] if len(cnt) > 1 else weekday_sig

            def pick_fill(cells2, sig):
                for cell in cells2:
                    if fill_sig_from_cell(cell) == sig:
                        return copy(cell.fill)
                return None

            weekday_fill = pick_fill(cells, weekday_sig)
            weekend_fill = pick_fill(cells, weekend_sig) or weekday_fill
            return weekday_fill, weekend_fill

        # Header örnek fill'leri
        header_cells = []
        for d in range(1, MAX_DAYS + 1):
            c = FIRST_DAY_COL + (d - 1)
            header_cells.append(ws.cell(HEADER_ROW, c))

        header_sigs = [fill_sig_from_cell(cell) for cell in header_cells]
        h_cnt = Counter(header_sigs).most_common()

        h_weekday_sig = h_cnt[0][0] if h_cnt else None
        h_weekend_sig = h_cnt[1][0] if len(h_cnt) > 1 else h_weekday_sig

        def pick_fill(cells, sig):
            for cell in cells:
                if fill_sig_from_cell(cell) == sig:
                    return copy(cell.fill)  # proxy yerine kopya al
            return None

        header_weekday_fill = pick_fill(header_cells, h_weekday_sig)
        header_weekend_fill = pick_fill(header_cells, h_weekend_sig) or header_weekday_fill

        # DT için örnek satır: 07:00 (row_idx=0)
        dt_row = GRID_START_ROW + 0
        # ODT için örnek satır: 11:00 (row_idx=16)  -> 10:15 sonrası ODT
        odt_row = GRID_START_ROW + 16

        dt_weekday_fill, dt_weekend_fill = analyze_grid_row(dt_row)
        odt_weekday_fill, odt_weekend_fill = analyze_grid_row(odt_row)

        disabled_fill = odt_weekday_fill or dt_weekday_fill

        for day in range(1, MAX_DAYS + 1):
            col = FIRST_DAY_COL + (day - 1)

            if day <= days_in_month:
                dow = TR_DOW[date(year, month, day).weekday()]
                ws.cell(HEADER_ROW, col).value = f"{dow}\n{day}"

                is_weekend = dow in ("Ct", "Pa")
                hf = header_weekend_fill if is_weekend else header_weekday_fill
            else:
                ws.cell(HEADER_ROW, col).value = None
                hf = header_weekday_fill  # header çok bozmasın
                gf = disabled_fill        # grid'i gri yap

            # Header fill uygula
            if hf is not None:
                ws.cell(HEADER_ROW, col).fill = hf

            # Grid fill uygula (C8..AG59) - satır tipine göre DT/ODT tonu koru
            for row_idx in range(0, GRID_ROWS):
                rr = GRID_START_ROW + row_idx

                if day > days_in_month:
                    if disabled_fill is not None:
                        ws.cell(rr, col).fill = disabled_fill
                    continue

                slot_type = classify_dt_odt(_row_idx_to_time(row_idx))
                if is_weekend:
                    gf2 = dt_weekend_fill if slot_type == "DT" else odt_weekend_fill
                else:
                    gf2 = dt_weekday_fill if slot_type == "DT" else odt_weekday_fill

                if gf2 is not None:
                    ws.cell(rr, col).fill = gf2

        # Bir önceki export'tan gizli kalmış olabilecek kolonlar sadece 29-31.
        # 1..28 gibi kolonlara dokunursak şablonun range genişlikleri bozulabiliyor.
        for day in range(29, 32):
            col_letter = get_column_letter(FIRST_DAY_COL + (day - 1))
            if col_letter in ws.column_dimensions:
                ws.column_dimensions[col_letter].hidden = False

        # Ay dışı gün kolonlarını gizle (29/30/31)
        for day in range(days_in_month + 1, MAX_DAYS + 1):
            col = FIRST_DAY_COL + (day - 1)
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].hidden = True
            # İstersen ekstra: genişliği de sıfıra yakın yap
            # ws.column_dimensions[col_letter].width = 0.1

    # Kanal adı (Excel'de T2)
    ch = str(payload.get("channel_name", "")).strip()
    if ch:
        ws["U2"].value = ch

    # Sabit başlık
    ws["V3"].value = "Rezervasyon Formu"

    # --- PLAN GRID: temizle + doldur ---
    plan_cells = payload.get("plan_cells") or {}

    # Template'te plan grid başlangıcı: C8 (gün 1) varsayımı
    # Satırlar: 07:00-20:00, 15 dk => 52 satır (8..59)
    # Kolonlar: gün 1..31 => C..AG (3..33)
    GRID_START_ROW = 8
    GRID_START_COL = 4   # D
    GRID_ROWS = 52
    GRID_DAYS_MAX = 31

    # 1) Önce tüm grid alanını boşalt (eski A'lar geri gelmesin diye)
    for r in range(GRID_START_ROW, GRID_START_ROW + GRID_ROWS):
        for c in range(GRID_START_COL, GRID_START_COL + GRID_DAYS_MAX):
            ws.cell(r, c).value = None

    # 2) Sonra uygulamada girilenleri bas
    # plan_cells formatı: {(row_idx, day): "A"} veya {"row_idx,day": "A"} gelebilir
    for key, val in plan_cells.items():
        try:
            if isinstance(key, str):
                # "12,5" gibi gelirse
                row_idx_str, day_str = key.split(",")
                row_idx = int(row_idx_str)
                day = int(day_str)
            else:
                row_idx, day = key  # tuple
                row_idx = int(row_idx)
                day = int(day)

            if not (0 <= row_idx < GRID_ROWS and 1 <= day <= GRID_DAYS_MAX):
                continue

            rr = GRID_START_ROW + row_idx
            cc = GRID_START_COL + (day - 1)
            ws.cell(rr, cc).value = str(val)

        except Exception:
            # bozuk key gelirse export'u patlatmayalım
            continue

    # --- Dinlenme Oranı (Kuşak yanında, yeni sütun B) ---
    # access_hour_map: {"07:00-08:00": 9.92, ...} (normalize edilmiş saat etiketleri)
    try:
        access_hour_map = payload.get("access_hour_map") or {}
        dinlenme_col = 2  # B
        for row_idx in range(GRID_ROWS):
            rr = GRID_START_ROW + row_idx
            tt = _row_idx_to_time(row_idx)
            hour_label = _norm_hour_label(f"{tt.hour:02d}:00-{(tt.hour+1):02d}:00")
            v = access_hour_map.get(hour_label)
            ws.cell(rr, dinlenme_col).value = "NA" if v is None else v
    except Exception:
        pass

    # --- Birim fiyatı (AO sütunu) ---
    # DT/ODT fiyatları: seçilen kanalın, plan tarihinin AY/YIL'ına göre.
    try:
        unit_price_col = column_index_from_string("AP")
        dt_price = float(payload.get("channel_price_dt") or 0)
        odt_price = float(payload.get("channel_price_odt") or 0)

        for row_idx in range(GRID_ROWS):
            rr = GRID_START_ROW + row_idx
            slot_type = classify_dt_odt(_row_idx_to_time(row_idx))
            ws.cell(rr, unit_price_col).value = dt_price if slot_type == "DT" else odt_price
    except Exception:
        # birim fiyatı basılamazsa export'u patlatma
        pass

    # --- Logo ---
    logo_path = template_path.parent / "RADIOSCOPE.PNG"
    if not logo_path.exists():
        logo_path = resource_path("assets/RADIOSCOPE.PNG")

    try:
        ws._images = []  # kalsın, üst üste binmesin
        if not logo_path.exists():
            raise FileNotFoundError(f"Logo bulunamadı: {logo_path}")

        img = Image(str(logo_path))
        img.width = 128
        img.height = 128

        ws.add_image(img, "AP1")

    except Exception as e:
        print("[DEBUG] logo add FAILED:", repr(e))
        pass

    # --- Alt alanlar (Kod listesi + süre hesabı) ---
    # Ajans komisyon oranı (%): şirket/ajansa göre değişebilir.
    try:
        rate = float(payload.get("agency_commission_pct"))
    except Exception:
        rate = 10.0
    try:
        ws["AN62"].value = f"Ajans Komisyonu % {int(rate) if float(rate).is_integer() else rate}"
        ws["AR62"].value = f"=(AR61*{rate}/100)"
    except Exception:
        pass

    # Kullanılan kodları grid'den topla (D8:AH59)
    used: dict[str, int] = {}
    for r in range(8, 60):
        for c in range(4, 4 + 31):
            v = ws.cell(row=r, column=c).value
            code = str(v or "").strip().upper()
            if not code:
                continue
            used[code] = used.get(code, 0) + 1

    # Kod tanımları map: code -> {sn, desc}
    code_defs = payload.get("code_defs") or []
    code_map: dict[str, dict] = {}
    for d in code_defs:
        try:
            c = str(d.get("code") or "").strip().upper()
            if not c:
                continue
            code_map[c] = {"sn": int(d.get("duration_sec") or 0), "desc": str(d.get("desc") or "").strip()}
        except Exception:
            continue
    if not code_map:
        sc = str(payload.get("spot_code") or "").strip().upper()
        if sc:
            code_map[sc] = {"sn": int(payload.get("spot_duration_sec") or 0), "desc": str(payload.get("code_definition") or "").strip()}

    # Kod tablosu alanını temizle ve doldur (A67: Kod, C67: Süre, D67: Açıklama)
    start_row = 67
    max_rows = 8
    for i in range(max_rows):
        rr = start_row + i
        for addr in (f"A{rr}", f"C{rr}", f"D{rr}", f"G{rr}"):
            try:
                ws[addr].value = None
            except Exception:
                pass

    ordered: list[str] = []
    for d in code_defs:
        c = str(d.get("code") or "").strip().upper()
        if c and c in used and c not in ordered:
            ordered.append(c)
    for c in sorted(used.keys()):
        if c not in ordered:
            ordered.append(c)

    ordered = ordered[:max_rows]
    for i, c in enumerate(ordered):
        rr = start_row + i
        dd = code_map.get(c) or {}
        ws[f"A{rr}"].value = c
        ws[f"C{rr}"].value = int(dd.get("sn") or 0)
        ws[f"D{rr}"].value = str(dd.get("desc") or "")

    # Toplam adet göstergesi (şablonda G67)
    try:
        ws["G67"].value = f"({sum(used.values())})" if used else "(0)"
    except Exception:
        pass

    # AM sütunu (Bedelli Süre): hücredeki koda göre saniye lookup
    if ordered:
        last_row = start_row + len(ordered) - 1
        tbl_rng = f"$A${start_row}:$C${last_row}"
        cols = [get_column_letter(4 + i) for i in range(31)]  # D..AH
        for r in range(8, 60):
            parts = [f"IFERROR(VLOOKUP({col}{r},{tbl_rng},3,FALSE),0)" for col in cols]
            ws[f"AM{r}"].value = f"=SUM({','.join(parts)})"
    else:
        for r in range(8, 60):
            ws[f"AM{r}"].value = 0

    # NOT: sabit, içerik değişebilir
    note = str(payload.get("note_text", "")).strip()
    ws["A77"].value = "NOT:" if not note else f"NOT: {note}"

    # İsim değişebilir
    if payload.get("prepared_by") is not None:
        ws["AK77"].value = str(payload.get("prepared_by", "")).strip()

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _export_excel_span_legacy(
    template_path: Path,
    out_path: Path,
    payload: dict[str, Any],
    month_matrices: dict[tuple[int, int], dict[str, str]],
    span_start: date,
    span_end: date,
    *,
    max_days_per_sheet: int = 31,
) -> Path:
    """Tek dosyada, seçili tarih aralığına göre (gün/ay bazlı) rezervasyon Excel çıktısı üretir.

    - UI span modunda grid tek tabloda görünür ama DB kayıtları ay ay bölünür.
    - Bu fonksiyon ay ay dosya üretmek yerine, aralığı *tek Excel dosyasında* sayfalara böler.
      (Bir sayfa en fazla 31 gün.)

    month_matrices formatı: {(year,month): {"row,day": "KOD", ...}, ...}
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Template bulunamadı: {template_path}")

    # normalize
    if span_start and span_end and span_start > span_end:
        span_start, span_end = span_end, span_start

    # build full date list
    dates: list[date] = []
    cur = span_start
    while cur <= span_end:
        dates.append(cur)
        cur = cur.fromordinal(cur.toordinal() + 1)

    wb = openpyxl.load_workbook(template_path)
    # Excel açılınca formüller yeniden hesaplasın
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    base_ws = _pick_worksheet(
        wb,
        "REZERVASYON ve PLANLAMA",
        contains_any=["rezervasyon", "planlama"],
    )

    # --- Span genelinde kullanılan kodlar / süre / tanımlar (A67/B67/D67 için) ---
    # Not: Span modunda grid tek tablo olsa da ay ay matrise bölünerek geliyor.
    # Excel şablonundaki formüller B67'ye bağlı olduğu için (saniye) burada doğru
    # şekilde doldurmak kritik.

    # code_defs: [{code, desc, duration_sec}, ...]
    _code_defs = payload.get("code_defs") or []
    code_map: dict[str, dict[str, Any]] = {}
    for d in _code_defs:
        try:
            c = str(d.get("code") or "").strip().upper()
            if not c:
                continue
            code_map[c] = {
                "desc": str(d.get("desc") or "").strip(),
                "duration_sec": float(d.get("duration_sec") or 0),
            }
        except Exception:
            continue

    # fallback: tekli alanlar
    fb_code = str(payload.get("spot_code") or "").strip().upper()
    if fb_code and fb_code not in code_map:
        code_map[fb_code] = {
            "desc": str(payload.get("code_definition") or "").strip(),
            "duration_sec": float(payload.get("spot_duration_sec") or 0),
        }

    # Span içindeki gerçek kullanım (kod adetleri)
    used_counts: dict[str, int] = {}
    for dte in dates:
        mm = month_matrices.get((dte.year, dte.month)) or {}
        for row_idx in range(52):
            v = mm.get(f"{row_idx},{dte.day}", "")
            c = str(v or "").strip().upper()
            if not c:
                continue
            used_counts[c] = used_counts.get(c, 0) + 1

    used_codes_sorted = sorted(used_counts.keys())

    # A67: kodlar
    a67_codes = ",".join(used_codes_sorted) if used_codes_sorted else str(payload.get("spot_code") or "").strip()

    # D67: tanımlar
    d67_descs: list[str] = []
    for c in used_codes_sorted:
        desc = str((code_map.get(c) or {}).get("desc") or "").strip()
        if desc:
            d67_descs.append(desc)
    d67_text = ",".join(d67_descs) if d67_descs else str(payload.get("code_definition") or "").strip()

    # B67: tek kodsa süre, çokluysa ağırlıklı ortalama
    b67_val: float | int = 0
    if len(used_codes_sorted) == 1:
        b67_val = float((code_map.get(used_codes_sorted[0]) or {}).get("duration_sec") or 0)
    elif len(used_codes_sorted) > 1:
        total = sum(used_counts.values())
        if total > 0:
            wsum = 0.0
            for c in used_codes_sorted:
                dur = float((code_map.get(c) or {}).get("duration_sec") or 0)
                wsum += float(used_counts.get(c, 0)) * dur
            b67_val = wsum / float(total)
        else:
            b67_val = 0
    else:
        # hiç hücre yoksa payload'tan
        b67_val = float(payload.get("spot_duration_sec") or 0)

    # Tam sayı ise int bas (Excel daha temiz)
    try:
        if abs(float(b67_val) - int(float(b67_val))) < 1e-9:
            b67_val = int(float(b67_val))
    except Exception:
        pass

    # AK77: prepared_by -> timestamp ekle
    pb = str(payload.get("prepared_by") or "").strip()
    if pb:
        # Eğer kullanıcı sadece isim girdiyse, sonuna tarih-saat ekle.
        # (UI confirm akışında zaten ekleniyor; span export UI'den direkt geliyor.)
        if " - " not in pb or len(pb.split(" - ")) == 1:
            pb = f"{pb} - {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    else:
        pb = datetime.now().strftime("%d.%m.%Y %H:%M")

    # Helper: apply header + fills for an arbitrary date chunk
    def apply_span_headers(ws, dates_chunk: list[date]) -> None:
        HEADER_ROW = 7
        FIRST_DAY_COL = 4  # D
        MAX_DAYS = 31
        GRID_START_ROW = 8
        GRID_ROWS = 52

        # --- Şablondan weekday/weekend fill örneklerini otomatik yakala (mevcut mantık) ---
        def fill_sig_from_cell(cell):
            f = cell.fill
            pt = getattr(f, "patternType", None)
            fg = getattr(getattr(f, "fgColor", None), "rgb", None)
            bg = getattr(getattr(f, "bgColor", None), "rgb", None)
            return (pt, fg, bg)

        def pick_fill(cells, sig):
            for cell in cells:
                if fill_sig_from_cell(cell) == sig:
                    return copy(cell.fill)
            return None

        def analyze_grid_row(sample_row: int):
            cells = []
            for d in range(1, MAX_DAYS + 1):
                c = FIRST_DAY_COL + (d - 1)
                cells.append(ws.cell(sample_row, c))
            sigs = [fill_sig_from_cell(cell) for cell in cells]
            cnt = Counter(sigs).most_common()
            weekday_sig = cnt[0][0] if cnt else None
            weekend_sig = cnt[1][0] if len(cnt) > 1 else weekday_sig
            weekday_fill = pick_fill(cells, weekday_sig)
            weekend_fill = pick_fill(cells, weekend_sig) or weekday_fill
            return weekday_fill, weekend_fill

        # Header örnek fill'leri
        header_cells = []
        for d in range(1, MAX_DAYS + 1):
            c = FIRST_DAY_COL + (d - 1)
            header_cells.append(ws.cell(HEADER_ROW, c))

        header_sigs = [fill_sig_from_cell(cell) for cell in header_cells]
        h_cnt = Counter(header_sigs).most_common()
        h_weekday_sig = h_cnt[0][0] if h_cnt else None
        h_weekend_sig = h_cnt[1][0] if len(h_cnt) > 1 else h_weekday_sig

        header_weekday_fill = pick_fill(header_cells, h_weekday_sig)
        header_weekend_fill = pick_fill(header_cells, h_weekend_sig) or header_weekday_fill

        # DT/ODT satır örnekleri
        dt_row = GRID_START_ROW + 0
        odt_row = GRID_START_ROW + 16
        dt_weekday_fill, dt_weekend_fill = analyze_grid_row(dt_row)
        odt_weekday_fill, odt_weekend_fill = analyze_grid_row(odt_row)
        disabled_fill = odt_weekday_fill or dt_weekday_fill

        # Önce tüm gün kolonlarını açık yap (önceki export gizlemiş olabilir)
        for day in range(1, MAX_DAYS + 1):
            col_letter = get_column_letter(FIRST_DAY_COL + (day - 1))
            if col_letter in ws.column_dimensions:
                ws.column_dimensions[col_letter].hidden = False

        # Kolonlara tarihleri bas
        for idx in range(1, MAX_DAYS + 1):
            col = FIRST_DAY_COL + (idx - 1)
            cell_h = ws.cell(HEADER_ROW, col)

            if idx <= len(dates_chunk):
                dte = dates_chunk[idx - 1]
                dow = TR_DOW[dte.weekday()]
                cell_h.value = f"{dow}\n{dte.day:02d}.{dte.month:02d}"
                is_weekend = dow in ("Ct", "Pa")
                hf = header_weekend_fill if is_weekend else header_weekday_fill
                if hf is not None:
                    cell_h.fill = hf

                # grid fill
                for row_idx in range(0, GRID_ROWS):
                    rr = GRID_START_ROW + row_idx
                    slot_type = classify_dt_odt(_row_idx_to_time(row_idx))
                    if is_weekend:
                        gf2 = dt_weekend_fill if slot_type == "DT" else odt_weekend_fill
                    else:
                        gf2 = dt_weekday_fill if slot_type == "DT" else odt_weekday_fill
                    if gf2 is not None:
                        ws.cell(rr, col).fill = gf2

            else:
                # chunk dışı kolon: gizle
                cell_h.value = None
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].hidden = True
                # grid'i de "disabled" tona çek
                if disabled_fill is not None:
                    for row_idx in range(0, GRID_ROWS):
                        rr = GRID_START_ROW + row_idx
                        ws.cell(rr, col).fill = disabled_fill

    # chunk'lere böl
    chunks: list[list[date]] = []
    for i in range(0, len(dates), max_days_per_sheet):
        chunks.append(dates[i : i + max_days_per_sheet])

    period_text = f"{span_start.day:02d}.{span_start.month:02d}.{span_start.year} - {span_end.day:02d}.{span_end.month:02d}.{span_end.year}"

    # İlk sheet: mevcut
    sheets: list[Any] = [base_ws]
    # Diğer sayfaları kopyala
    for _ in range(1, len(chunks)):
        ws2 = wb.copy_worksheet(base_ws)
        sheets.append(ws2)

    # Sayfa isimleri
    for i, ws in enumerate(sheets, start=1):
        if len(chunks) == 1:
            ws.title = "REZERVASYON"
        else:
            c0 = chunks[i - 1][0]
            c1 = chunks[i - 1][-1]
            ws.title = f"{c0.day:02d}.{c0.month:02d}-{c1.day:02d}.{c1.month:02d}"

    # Her sheet'i doldur
    for si, ws in enumerate(sheets):
        dates_chunk = chunks[si]

        # --- Header labels + values (aynı) ---
        ws["A1"].value = "Ajans:"
        ws["C1"].value = str(payload.get("agency_name", "")).strip()

        ws["A2"].value = "Reklam Veren:"
        ws["C2"].value = str(payload.get("advertiser_name", "")).strip()

        ws["A3"].value = "Ürün:"
        ws["C3"].value = str(payload.get("product_name", "")).strip()

        ws["A4"].value = "Plan Başlığı:"
        ws["C4"].value = str(payload.get("plan_title", "")).strip()

        ws["A5"].value = "Rezervasyon No:"
        ws["C5"].value = str(payload.get("reservation_no", "")).strip()

        hdr_font = copy(ws["A1"].font)
        for addr in ("A4", "A5", "A6"):
            ws[addr].font = hdr_font

        ws["A6"].value = "Dönemi:"
        ws["C6"].value = period_text

        # Kanal adı
        ch = str(payload.get("channel_name", "")).strip()
        if ch:
            ws["U2"].value = ch
            # Şablonda T2 kanal adı fontu 13 olmalı (bazı akışlarda 11'e düşüyordu)
            try:
                ws["U2"].font = copy(ws["U2"].font)
                ws["U2"].font = ws["U2"].font.copy(size=13)
            except Exception:
                pass

        ws["V3"].value = "Rezervasyon Formu"

        # Span tarih başlıkları + boyamalar
        apply_span_headers(ws, dates_chunk)

        # --- PLAN GRID: temizle + doldur ---
        GRID_START_ROW = 8
        GRID_START_COL = 4  # D (C is Dolar Kuru)
        GRID_ROWS = 52

        # temizle: sadece chunk uzunluğu kadar kolon açık; ama şablon 31 kolon olduğu için 31'i temizleyelim
        for r in range(GRID_START_ROW, GRID_START_ROW + GRID_ROWS):
            for c in range(GRID_START_COL, GRID_START_COL + 31):
                ws.cell(r, c).value = None

        for row_idx in range(GRID_ROWS):
            for di, dte in enumerate(dates_chunk):
                mm = month_matrices.get((dte.year, dte.month)) or {}
                code = mm.get(f"{row_idx},{dte.day}", "")
                if not str(code).strip():
                    continue
                rr = GRID_START_ROW + row_idx
                cc = GRID_START_COL + di
                ws.cell(rr, cc).value = str(code)

        # --- Dinlenme Oranı (Kuşak yanında, yeni sütun B) ---
        # access_hour_map: {"07:00-08:00": 9.92, ...} (normalize edilmiş saat etiketleri)
        try:
            access_hour_map = payload.get("access_hour_map") or {}
            dinlenme_col = 2  # B
            dolar_col = 3  # C
            dolar_kuru = payload.get('dolar_kuru', payload.get('dolar', 2))
            for row_idx in range(GRID_ROWS):
                rr = GRID_START_ROW + row_idx
                tt = _row_idx_to_time(row_idx)
                hour_label = _norm_hour_label(f"{tt.hour:02d}:00-{(tt.hour+1):02d}:00")
                v = access_hour_map.get(hour_label)
                ws.cell(rr, dinlenme_col).value = "NA" if v is None else v
                ws.cell(rr, dolar_col).value = dolar_kuru
        except Exception:
            pass

        # --- Birim fiyatı (AP sütunu) ---
        try:
            unit_price_col = column_index_from_string("AP")
            dt_price = float(payload.get("channel_price_dt") or 0)
            odt_price = float(payload.get("channel_price_odt") or 0)
            for row_idx in range(GRID_ROWS):
                rr = GRID_START_ROW + row_idx
                slot_type = classify_dt_odt(_row_idx_to_time(row_idx))
                ws.cell(rr, unit_price_col).value = dt_price if slot_type == "DT" else odt_price
        except Exception:
            pass

        # --- Logo (her sheet için) ---
        logo_path = template_path.parent / "RADIOSCOPE.PNG"
        if not logo_path.exists():
            logo_path = resource_path("assets/RADIOSCOPE.PNG")
        try:
            ws._images = []
            if logo_path.exists():
                img = Image(str(logo_path))
                img.width = 128
                img.height = 128
                ws.add_image(img, "AP1")
        except Exception:
            pass

        # --- Alt alanlar ---
        # A67/B67/D67: Span genelindeki gerçek seçimlere göre doldur.
        if a67_codes is not None:
            ws["A67"].value = str(a67_codes).strip()

        try:
            ws["C67"].value = b67_val
        except Exception:
            pass

        # toplam adet: sadece bu chunk içindeki dolu hücre sayısı
        adet_total = 0
        for row_idx in range(GRID_ROWS):
            for dte in dates_chunk:
                mm = month_matrices.get((dte.year, dte.month)) or {}
                v = mm.get(f"{row_idx},{dte.day}", "")
                if str(v).strip():
                    adet_total += 1
        ws["F67"].value = f"({int(adet_total)})"

        ws["D67"].value = d67_text if str(d67_text).strip() else None

        note = str(payload.get("note_text", "")).strip()
        ws["A77"].value = "NOT:" if not note else f"NOT: {note}"

        ws["AK77"].value = pb

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path





def export_excel_span(template_path: Path, out_path: Path, payload: dict, month_matrices: dict, span_start: date, span_end: date) -> None:
    """
    Export a reservation form workbook for an arbitrary date span.

    New template (v2) rules (based on user's reference workbook):
      - Column A: Kuşak (already in template)
      - Column B: Dinlenme Oranı (hourly access ratio; repeated for 15-min slots)
      - Column C: Dolar Kuru
      - Day grid starts at column D and spans max 31 days (D..AH)
      - Totals columns are fixed (AI..AR) and rely on formulas in the template.

    Notes:
      - We NEVER insert/delete columns. We only write into the existing template cells.
      - Template images are kept for the first sheet; for copied sheets we re-add the logo.
    """
    template_path = Path(template_path)

    # If caller gave the old template, but we have the new one shipped with the app, prefer it.
    # (This prevents "column shift" bugs when Dinlenme Oranı is required.)
    try:
        shipped_v2 = Path(__file__).resolve().parents[2] / "assets" / "reservation_template.xlsx"
        if shipped_v2.exists():
            template_path = shipped_v2
    except Exception:
        pass

    # NOTE: use the already-imported openpyxl module. Some earlier patches
    # accidentally used load_workbook without importing it, which crashes
    # the export path at runtime.
    wb = openpyxl.load_workbook(template_path)

    # Pick template sheet
    if "TEMPLATE" in wb.sheetnames:
        ws_tmpl = wb["TEMPLATE"]
    else:
        ws_tmpl = wb[wb.sheetnames[0]]
        ws_tmpl.title = "TEMPLATE"

    # Detect v2 template by header row
    is_v2 = (ws_tmpl["A7"].value == "Kuşak" and str(ws_tmpl["B7"].value).strip().lower().startswith("dinlenme"))

    if not is_v2:
        # Fallback to legacy exporter (kept for backward compatibility)
        _export_excel_span_legacy(template_path=template_path, out_path=out_path, payload=payload, month_matrices=month_matrices, span_start=span_start, span_end=span_end)
        return

    # Keep only the template sheet in workbook (clean output)
    for sn in list(wb.sheetnames):
        if sn != ws_tmpl.title:
            del wb[sn]

    # Extract logo from template xlsx (copy_worksheet doesn't copy images)
    logo_bytes = None
    logo_anchor = "AP1"
    logo_w = logo_h = None
    try:
        if getattr(ws_tmpl, "_images", None):
            img0 = ws_tmpl._images[0]
            logo_w = getattr(img0, "width", None)
            logo_h = getattr(img0, "height", None)
            try:
                col0 = img0.anchor._from.col  # 0-based
                row0 = img0.anchor._from.row  # 0-based
                logo_anchor = f"{get_column_letter(col0+1)}{row0+1}"
            except Exception:
                pass

        import zipfile
        with zipfile.ZipFile(template_path, "r") as zf:
            media = sorted([n for n in zf.namelist() if n.startswith("xl/media/") and n.lower().endswith((".png", ".jpg", ".jpeg"))])
            if media:
                logo_bytes = zf.read(media[0])
    except Exception:
        logo_bytes = None

    def _add_logo(ws):
        if not logo_bytes:
            return
        try:
            from io import BytesIO
            img = Image(BytesIO(logo_bytes))
            if logo_w and logo_h:
                img.width = logo_w
                img.height = logo_h
            ws.add_image(img, logo_anchor)
        except Exception:
            # No hard fail: export should still work without logo.
            pass

    # Build list of dates inclusive, chunked by calendar months (for workbook sheets)
    # Example: 2026-03-09..2026-04-13 => [09.03-31.03] + [01.04-13.04]
    chunks: list[list[date]] = []
    cur = span_start
    while cur <= span_end:
        last_dom = calendar.monthrange(cur.year, cur.month)[1]
        month_end = date(cur.year, cur.month, last_dom)
        seg_end = month_end if month_end <= span_end else span_end

        chunk: list[date] = []
        d0 = cur
        while d0 <= seg_end:
            chunk.append(d0)
            d0 += timedelta(days=1)

        chunks.append(chunk)
        cur = seg_end + timedelta(days=1)

    # Normalize access_hour_map keys once
    access_hour_map = payload.get("access_hour_map") or {}
    norm_access = {}
    for k, v in access_hour_map.items():
        if not k:
            continue
        kk = str(k).replace(" ", "")
        # Normalize "07:00-08:00" / "7-8" etc to "07:00-08:00"
        m = re.match(r"^(\d{1,2})(?::\d{2})?-(\d{1,2})(?::\d{2})?$", kk)
        if m:
            h1 = int(m.group(1)); h2 = int(m.group(2))
            kk = f"{h1:02d}:00-{h2:02d}:00"
        norm_access[kk] = v

    def _dow_tr(dt_: date) -> str:
        return ["Pt", "Sa", "Ça", "Pe", "Cu", "Ct", "Pa"][dt_.weekday()]

    def _fill_header_and_grid(ws, dates_chunk):
        # Clear header and grid first
        for idx in range(31):
            col = 4 + idx  # D=4
            ws.cell(row=7, column=col).value = None
            for r in range(8, 60):  # 8..59
                ws.cell(row=r, column=col).value = None

        # Write headers
        for idx, dt_ in enumerate(dates_chunk):
            col = 4 + idx
            hcell = ws.cell(row=7, column=col)
            hcell.value = f"{_dow_tr(dt_)}\n{dt_:%d.%m}"
            # keep style; ensure wrap
            try:
                hcell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            except Exception:
                pass

    def _get(payload, *keys, default=""):
        for k in keys:
            v = payload.get(k)
            if v is not None and str(v).strip() != "":
                return v
        return default

    def _fill_meta(ws, span_start, span_end):
        ws["B1"] = _get(payload, "agency_name", "agency")
        ws["B2"] = _get(payload, "advertiser_name", "advertiser")
        ws["B3"] = _get(payload, "product_name", "product")
        ws["B4"] = _get(payload, "plan_title")
        ws["B5"] = _get(payload, "reservation_no")

        period_txt = f"{span_start:%d.%m.%Y} - {span_end:%d.%m.%Y}"
        # Şablonda dönem B6'dadır. D6'ya yazınca çıktı içinde "dönem" iki kere görünüyor.
        ws["B6"] = period_txt
        try:
            ws["D6"].value = None
        except Exception:
            pass

        ch = _get(payload, "channel_name", "channel")
        if ch:
            ws["U2"] = ch

        note = _get(payload, "note_text", "note")
        if note:
            note_str = str(note).strip()
            ws["A77"] = note_str if note_str.lower().startswith("not") else f"NOT: {note_str}"


    def _fill_rates(ws):
        usd = payload.get("usd_rate")
        if usd is None:
            usd = payload.get("dolar_kuru", payload.get("dolar", None))
        if usd is None:
            usd = 2  # default as per your examples
        for r in range(8, 60):
            ws[f"C{r}"].value = usd

        # Dinlenme oranı per slot (hourly repeated)
        for idx in range(0, 52):  # 52 quarter-hour rows
            excel_row = 8 + idx
            hour = 7 + (idx // 4)  # 07:00 starts at row 8
            key = f"{hour:02d}:00-{hour+1:02d}:00"
            val = norm_access.get(key, "NA")
            ws[f"B{excel_row}"].value = val

    def _fill_codes(ws, dates_chunk):
        for row_idx in range(0, 52):  # planning grid rows
            excel_row = 8 + row_idx
            for col_idx, dt_ in enumerate(dates_chunk):
                # get code from month_matrices
                mm = month_matrices.get((dt_.year, dt_.month), {}) or {}
                code = mm.get(f"{row_idx},{dt_.day}", "")
                ws.cell(row=excel_row, column=4 + col_idx).value = (str(code).strip().upper() if str(code).strip() else "")

    def _apply_commission(ws):
        # Ajans komisyon oranı değişken: %0/%5/%10/... (UI'dan gelecek)
        rate = payload.get("agency_commission_pct")
        try:
            rate = float(rate)
        except Exception:
            rate = 10.0
        # Etiket + formül
        try:
            ws["AN62"].value = f"Ajans Komisyonu % {int(rate) if float(rate).is_integer() else rate}"
        except Exception:
            pass
        try:
            ws["AR62"].value = f"=(AR61*{rate}/100)"
        except Exception:
            pass

    def _fill_code_table_and_duration_formulas(ws):
        """Kod tablosunu (A67:C?) doldur ve AM8:AM59 'Bedelli Süre' formüllerini
        hücredeki koda göre saniye lookup yapacak şekilde güncelle.

        Şablondaki eski mantık: AMr = AIr * $C$67 (tek bir ortalama süre)
        Yeni mantık: AMr = SUM( VLOOKUP(D..AH, kod->sn tablosu) )
        """
        # 1) Grid'deki kullanılan kodları topla
        used: dict[str, int] = {}
        for r in range(8, 60):
            for c in range(4, 4 + 31):  # D..AH
                v = ws.cell(row=r, column=c).value
                code = str(v or "").strip().upper()
                if not code:
                    continue
                used[code] = used.get(code, 0) + 1

        # 2) Kod tanımlarından map oluştur (desc + sn)
        code_defs = payload.get("code_defs") or []
        code_map: dict[str, dict] = {}
        for d in code_defs:
            try:
                c = str(d.get("code") or "").strip().upper()
                if not c:
                    continue
                code_map[c] = {
                    "sn": int(d.get("duration_sec") or 0),
                    "desc": str(d.get("desc") or "").strip(),
                }
            except Exception:
                continue
        # geri uyumluluk
        if not code_map:
            sc = str(payload.get("spot_code") or "").strip().upper()
            if sc:
                code_map[sc] = {
                    "sn": int(payload.get("spot_duration_sec") or 0),
                    "desc": str(payload.get("code_definition") or "").strip(),
                }

        # 3) Tabloyu doldur (A67: Kod, C67: Süre, D67: Açıklama)
        start_row = 67
        max_rows = 8  # şablonda alan kısıtlı; gerekirse artırırız

        # önce temizle
        for i in range(max_rows):
            rr = start_row + i
            for addr in (f"A{rr}", f"C{rr}", f"D{rr}", f"G{rr}"):
                try:
                    ws[addr].value = None
                except Exception:
                    pass

        # kod sırası: code_defs sırası (varsa) + kalanlar alfabetik
        ordered: list[str] = []
        for d in code_defs:
            c = str(d.get("code") or "").strip().upper()
            if c and c in used and c not in ordered:
                ordered.append(c)
        for c in sorted(used.keys()):
            if c not in ordered:
                ordered.append(c)

        ordered = ordered[:max_rows]
        for i, c in enumerate(ordered):
            rr = start_row + i
            dd = code_map.get(c) or {}
            ws[f"A{rr}"].value = c
            ws[f"C{rr}"].value = int(dd.get("sn") or 0)
            # açıklama hücreleri şablonda D..F merge; sadece D'ye yazmak yeter
            ws[f"D{rr}"].value = str(dd.get("desc") or "")

        # toplam adet göstergesi (G67) (şablonda parantezli)
        try:
            ws["G67"].value = f"({sum(used.values())})" if used else "(0)"
        except Exception:
            pass

        # 4) AM sütunu formülleri: her hücre koda göre saniye bulsun.
        if ordered:
            last_row = start_row + len(ordered) - 1
            tbl_rng = f"$A${start_row}:$C${last_row}"
            cols = [get_column_letter(4 + i) for i in range(31)]  # D..AH
            for r in range(8, 60):
                parts = [f"IFERROR(VLOOKUP({col}{r},{tbl_rng},3,FALSE),0)" for col in cols]
                ws[f"AM{r}"].value = f"=SUM({','.join(parts)})"
        else:
            for r in range(8, 60):
                ws[f"AM{r}"].value = 0

    # Build output workbook sheets
    # First chunk uses TEMPLATE sheet; others are copies (and we re-add logo)
    for ci, chunk in enumerate(chunks):
        ch_start, ch_end = chunk[0], chunk[-1]
        title = f"{ch_start:%d.%m}-{ch_end:%d.%m}"

        if ci == 0:
            ws = ws_tmpl
        else:
            ws = wb.copy_worksheet(ws_tmpl)
            _add_logo(ws)

        ws.title = title

        _fill_meta(ws, ch_start, ch_end)
        _fill_header_and_grid(ws, chunk)
        _fill_rates(ws)
        _fill_codes(ws, chunk)
        _apply_commission(ws)
        _fill_code_table_and_duration_formulas(ws)

    # Save
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def export_plan_ozet_yearly(out_path, data: dict) -> None:
    """
    Yıllık Plan Özet Excel çıktısı:
    - 1 adet TOPLAM sheet (12 ay sütunu + yıl toplamı)
    - 12 adet ay sheet (aylık plan özet formatı)

    data beklenen format:
    {
      "template_path": "...",
      "year": 2026,
      "total": {"header": {...}, "rows": [...], "month_labels": [...]},
      "months": [ {"header": {...}, "rows":[...], "month": 3}, ... ]  # 12 eleman
    }
    """
    out_path = Path(out_path)

    template_path = data.get("template_path")
    if not template_path:
        template_path = Path(__file__).resolve().parents[2] / "assets" / "plan_ozet_template.xlsx"
    template_path = Path(template_path)

    wb = openpyxl.load_workbook(template_path)
    ws_template = wb.active  # şablonda tek sheet var

    def _fill_plan_ozet_sheet(ws, header: dict, rows: list[dict], day_labels: list[str], period_text: str):
        # Sabit hücreler
        ws["C2"].value = header.get("agency", "")
        ws["C3"].value = header.get("advertiser", "")
        ws["C4"].value = header.get("product", "")
        ws["C5"].value = header.get("plan_title", "")
        ws["C6"].value = header.get("reservation_no", "")
        ws["C7"].value = period_text
        ws["C8"].value = header.get("spot_len", 0)

        start_row = 11

        col_channel = 2     # B
        col_group = 3       # C
        col_dtodt = 4       # D
        col_ratio = 5       # E
        day_start_col = 6   # F
        col_month_adet = 37 # AK
        col_month_sn = 38   # AL
        col_unit_price = 39 # AM
        col_budget = 40     # AN

        # Header row (gün/ay)
        header_row = 10
        for i in range(31):
            ws.cell(header_row, day_start_col + i).value = None
        for i, lbl in enumerate(day_labels[:31]):
            ws.cell(header_row, day_start_col + i).value = lbl

        # Eski veriyi temizle (500 satır yeter)
        for rr in range(start_row, start_row + 500):
            for cc in range(col_channel, col_budget + 1):
                ws.cell(rr, cc).value = None

        r = start_row
        for row in rows:
            ws.cell(r, col_channel).value = row.get("channel", "")
            ws.cell(r, col_group).value = row.get("group", "")
            ws.cell(r, col_dtodt).value = row.get("dt_odt", "")
            ws.cell(r, col_ratio).value = row.get("ratio", "NA")

            days = row.get("days", []) or []
            for i in range(31):
                val = days[i] if i < len(days) else None
                ws.cell(r, day_start_col + i).value = val if val not in ("", None) else None

            unit = row.get("unit_price", 0)
            try:
                ws.cell(r, col_unit_price).value = float(unit)
            except Exception:
                ws.cell(r, col_unit_price).value = 0.0

            # Formüller (Excel açılınca hesaplanır)
            ws.cell(r, col_month_adet).value = f"=SUM(F{r}:AJ{r})"
            ws.cell(r, col_month_sn).value = f"=AK{r}*$C$8"
            ws.cell(r, col_budget).value = f"=AL{r}*AM{r}"

            r += 1

        last_data_row = r - 1
        total_row = last_data_row + 1
        ws.cell(total_row, col_channel).value = "Toplam"

        if last_data_row >= start_row:
            for i in range(31):
                c = day_start_col + i
                col_letter = get_column_letter(c)
                ws.cell(total_row, c).value = f"=SUM({col_letter}{start_row}:{col_letter}{last_data_row})"

            ws.cell(total_row, col_month_adet).value = f"=SUM(AK{start_row}:AK{last_data_row})"
            ws.cell(total_row, col_month_sn).value = f"=SUM(AL{start_row}:AL{last_data_row})"
            ws.cell(total_row, col_budget).value = f"=SUM(AN{start_row}:AN{last_data_row})"
        else:
            ws.cell(total_row, col_month_adet).value = 0
            ws.cell(total_row, col_month_sn).value = 0
            ws.cell(total_row, col_budget).value = 0

    # TOPLAM sheet
    total_block = data.get("total") or {}
    total_header = total_block.get("header") or {}
    total_rows = total_block.get("rows") or []
    month_labels = total_block.get("month_labels") or [
        "OCAK","ŞUBAT","MART","NİSAN","MAYIS","HAZİRAN","TEMMUZ","AĞUSTOS","EYLÜL","EKİM","KASIM","ARALIK"
    ]

    ws_total = wb.copy_worksheet(ws_template)
    ws_total.title = "TOPLAM"

    year = total_header.get("year") or data.get("year") or ""
    _fill_plan_ozet_sheet(
        ws_total,
        header={
            "agency": total_header.get("agency", ""),
            "advertiser": total_header.get("advertiser", ""),
            "product": total_header.get("product", ""),
            "plan_title": total_header.get("plan_title", ""),
            "reservation_no": "",
            "spot_len": total_header.get("spot_len", 0),
        },
        rows=total_rows,
        day_labels=month_labels,
        period_text=f"{year} (TOPLAM)"
    )

    # 12 ay sheet
    months = data.get("months") or []
    month_names = ["OCAK","ŞUBAT","MART","NİSAN","MAYIS","HAZİRAN","TEMMUZ","AĞUSTOS","EYLÜL","EKİM","KASIM","ARALIK"]

    for i in range(12):
        mdata = months[i] if i < len(months) else {"header": total_header, "rows": [], "month": i + 1}
        header = mdata.get("header") or {}
        rows = mdata.get("rows") or []
        month_no = int(mdata.get("month") or (i + 1))

        ws_m = wb.copy_worksheet(ws_template)
        ws_m.title = f"{month_no:02d}-{month_names[i]}"

        day_labels = [str(d) for d in range(1, 32)]
        period_txt = header.get("period") or f"{month_names[i]} {header.get('year', year)}"
        _fill_plan_ozet_sheet(
            ws_m,
            header={
                "agency": header.get("agency", ""),
                "advertiser": header.get("advertiser", ""),
                "product": header.get("product", ""),
                "plan_title": header.get("plan_title", ""),
                "reservation_no": "",
                "spot_len": header.get("spot_len", total_header.get("spot_len", 0)),
            },
            rows=rows,
            day_labels=day_labels,
            period_text=period_txt
        )

    wb.remove(ws_template)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def export_kod_tanimi(out_path, advertiser_name: str, rows: list[dict]) -> None:
    """
    KOD TANIMI çıktısını, assets/kod_tanimi_template.xlsx şablonunun görsel stilini koruyarak üretir.
    Tablo yerleşimi şablondaki gibi C5:F13 aralığındadır.
    """
    out_path = Path(out_path)

    template_path = resource_path("assets/kod_tanimi_template.xlsx")
    if template_path.exists():
        wb = openpyxl.load_workbook(template_path)
        ws = wb["KOPYA TANIMI"] if "KOPYA TANIMI" in wb.sheetnames else wb.active
    else:
        # Fallback: basit boş dosya
        wb = Workbook()
        ws = wb.active
        ws.title = "KOPYA TANIMI"
        # Header
        ws["C5"].value = "Kod"
        ws["D5"].value = "Kod Tanımı"
        ws["E5"].value = "Kod Uzunluğu (SN)"
        ws["F5"].value = "Dağılım"

    start_row = 6
    base_rows = 7  # şablonda 6..12
    total_row = start_row + base_rows  # 13

    # --- satır sayısını ihtiyaca göre büyüt (7'yi aşarsa) ---
    n = len(rows)
    if n > base_rows:
        insert_count = n - base_rows
        ws.insert_rows(total_row, amount=insert_count)

        # 12. satırın (son veri satırı) stilini yeni satırlara kopyala
        src_row = total_row - 1  # eski 12
        for i in range(insert_count):
            dst_row = src_row + 1 + i
            for col in range(3, 7):  # C..F
                src = ws.cell(src_row, col)
                dst = ws.cell(dst_row, col)
                dst._style = copy(src._style)
                dst.number_format = src.number_format
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.alignment = copy(src.alignment)
                dst.protection = copy(src.protection)
                dst.comment = None

        total_row += insert_count

    # --- önce tabloyu temizle (şablondaki örnek veriler varsa sil) ---
    max_data_rows = max(base_rows, n)
    for r in range(start_row, start_row + max_data_rows):
        for c in range(3, 7):  # C..F
            ws.cell(r, c).value = None

    # --- verileri yaz ---
    for i, r in enumerate(rows):
        rr = start_row + i
        ws.cell(rr, 3).value = r.get("code", "")
        ws.cell(rr, 4).value = r.get("code_desc", "")
        ws.cell(rr, 5).value = int(r.get("length_sn", 0) or 0)
        ws.cell(rr, 6).value = float(r.get("distribution", 0.0) or 0.0)

        # Dağılım yüzde formatında görünsün
        ws.cell(rr, 6).number_format = "0%"

    # --- kalan satırları boş bırak ama format kalsın ---
    # (şablon zaten formatlı, biz sadece value None bıraktık)

    # --- Ort.Uzun. formülleri: dinamik aralık ---
    last_data_row = start_row + max_data_rows - 1
    ws.cell(total_row, 3).value = "Ort.Uzun."
    ws.cell(total_row, 5).value = f"=SUMPRODUCT(E{start_row}:E{last_data_row},F{start_row}:F{last_data_row})"
    ws.cell(total_row, 6).value = f"=SUM(F{start_row}:F{last_data_row})"

    # toplam dağılım hücresi yüzde görünsün
    ws.cell(total_row, 6).number_format = "0%"

    # küçük bir başlık (dosya adı vs.) istersen buraya eklenir; şimdilik dokunmuyoruz.
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def export_spotlist(out_path, advertiser_name: str, rows: list[dict]) -> None:
    """SPOTLİST+ çıktısını şablonun stilini bozmadan üretir.

    Şablon: assets/spotlist_template.xlsx (sheet: "SPOTLİST+")
    Data satırları: 2. satırdan itibaren.
    """
    out_path = Path(out_path)

    template_path = resource_path("assets/spotlist_template.xlsx")
    if template_path.exists():
        wb = openpyxl.load_workbook(template_path)
        ws = wb["SPOTLİST+"] if "SPOTLİST+" in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "SPOTLİST+"
        headers = [
            "Sıra",
            "TARIH",
            "ANA YAYIN",
            "REKLAMIN FIRMASI",
            "ADET",
            "BASLANGIC",
            "SURE",
            "Spot Kodu",
            "DT-ODT",
            "Birim Saniye ",
            "Bütçe  Net TL",
            "Yayınlandı Durum",
        ]
        for i, h in enumerate(headers, start=1):
            ws.cell(1, i).value = h

    start_row = 2
    n = len(rows)

    # Şablonda hazır satır yoksa (veya yetmezse) satır ekle ve stil kopyala
    if ws.max_row < start_row:
        ws.insert_rows(start_row, amount=1)

    # Stil kopyalama için referans satır: start_row
    style_row = start_row if ws.max_row >= start_row else 1

    needed_last = start_row + max(n, 0) - 1
    if needed_last > ws.max_row:
        insert_count = needed_last - ws.max_row
        ws.insert_rows(ws.max_row + 1, amount=insert_count)
        # son mevcut stil satırını kopyala
        src_row = style_row
        for i in range(0, insert_count):
            dst_row = ws.max_row - insert_count + 1 + i
            for col in range(1, 13):
                src = ws.cell(src_row, col)
                dst = ws.cell(dst_row, col)
                dst._style = copy(src._style)
                dst.number_format = src.number_format
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.alignment = copy(src.alignment)
                dst.protection = copy(src.protection)
                dst.comment = None

    # Önce eski değerleri temizle (önceki export'tan kalmasın)
    # Template çok büyük olabiliyor, o yüzden sadece "kullanılmış" aralığı temizliyoruz.
    last_used = 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip() != "":
            last_used = r
    clear_last = max(last_used, needed_last)
    for r in range(start_row, clear_last + 1):
        for c in range(1, 13):
            ws.cell(r, c).value = None

    # Veriyi bas
    for i, rr in enumerate(rows):
        r = start_row + i
        ws.cell(r, 1).value = int(rr.get("sira", i + 1) or (i + 1))
        ws.cell(r, 2).value = rr.get("tarih", "")
        ws.cell(r, 3).value = rr.get("ana_yayin", "")
        ws.cell(r, 4).value = rr.get("reklam_firmasi", advertiser_name)
        ws.cell(r, 5).value = int(rr.get("adet", 1) or 1)
        ws.cell(r, 6).value = rr.get("baslangic", "")
        ws.cell(r, 7).value = int(rr.get("sure", 0) or 0)
        ws.cell(r, 8).value = rr.get("spot_kodu", "")
        ws.cell(r, 9).value = rr.get("dt_odt", "")
        ws.cell(r, 10).value = float(rr.get("birim_saniye", 0.0) or 0.0)
        ws.cell(r, 11).value = float(rr.get("butce_net", 0.0) or 0.0)
        ws.cell(r, 12).value = int(rr.get("published", 0) or 0)

    
    # Özet satırı (filtrelenmiş listeye göre)
    total_adet = sum(int(r.get("adet", 1) or 1) for r in rows) if rows else 0
    total_budget = sum(float(r.get("butce_net", 0.0) or 0.0) for r in rows) if rows else 0.0
    durations = [int(r.get("sure", 0) or 0) for r in rows] if rows else []
    avg_duration = (sum(durations) / len(durations)) if durations else 0.0

    sum_row = start_row + n
    if sum_row > ws.max_row:
        ws.insert_rows(ws.max_row + 1, amount=(sum_row - ws.max_row))

    # Stil kopyala (kenarlıklar vs kalsın)
    for col in range(1, 13):
        src = ws.cell(style_row, col)
        dst = ws.cell(sum_row, col)
        dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.comment = None

    # Değerleri yaz
    for c in range(1, 13):
        ws.cell(sum_row, c).value = None

    ws.cell(sum_row, 2).value = "TOPLAM"
    ws.cell(sum_row, 5).value = int(total_adet)
    ws.cell(sum_row, 7).value = float(avg_duration)
    ws.cell(sum_row, 11).value = float(total_budget)

    # Kalın yaz
    for c in range(1, 13):
        cell = ws.cell(sum_row, c)
        cell.font = copy(cell.font)
        cell.font = cell.font.copy(bold=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ------------------------------
# PLAN ÖZET (ay bazlı birleştirilmiş özet)
# ------------------------------

def export_plan_ozet(out_path, data: dict) -> None:
    """Plan Özet ekranındaki tabloyu, örnek PLAN ÖZET şablonuyla Excel'e döker.

    Not: Formüller openpyxl tarafından hesaplanmaz. Excel açıldığında hesaplanır.
    """

    def _copy_row_style(ws, src_row: int, dst_row: int, min_col: int, max_col: int) -> None:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
        for col in range(min_col, max_col + 1):
            s = ws.cell(src_row, col)
            d = ws.cell(dst_row, col)
            d._style = copy(s._style)
            d.number_format = s.number_format
            d.font = copy(s.font)
            d.border = copy(s.border)
            d.fill = copy(s.fill)
            d.alignment = copy(s.alignment)
            d.protection = copy(s.protection)
            d.comment = None

    out_path = Path(out_path)

    template_path = resource_path("assets/plan_ozet_template.xlsx")
    if not template_path.exists():
        raise FileNotFoundError(f"Plan özet şablonu bulunamadı: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb["PLAN ÖZET"] if "PLAN ÖZET" in wb.sheetnames else wb.active

    header = data.get("header") or {}
    rows = data.get("rows") or []
    days_in_month = int(data.get("days") or 31)

    # --- Üst bilgiler ---
    ws["C2"].value = str(header.get("agency", "") or "")
    ws["C3"].value = str(header.get("advertiser", "") or "")
    ws["C4"].value = str(header.get("product", "") or "")
    ws["C5"].value = str(header.get("plan_title", "") or "")
    ws["C6"].value = str(header.get("reservation_no", "") or "")
    ws["C7"].value = str(header.get("period", "") or "")

    spot_len = header.get("spot_len", 0) or 0
    try:
        ws["C8"].value = float(spot_len)
    except Exception:
        ws["C8"].value = 0

    month_name = str(header.get("month_name", "") or "").strip().upper()

    # --- Ay başlığı / kolon başlıkları ---
    if month_name:
        ws["F9"].value = month_name
        ws["AK9"].value = f"{month_name} Toplam"
        ws["AK10"].value = f"{month_name} Adet"
        ws["AL10"].value = f"{month_name} Saniye"

    # Gün başlıkları (1..31)
    header_row = 10
    day_start_col = 6   # F
    for d in range(1, 32):
        c = day_start_col + (d - 1)
        ws.cell(header_row, c).value = d if d <= days_in_month else None

    # --- Tablo alanı boyutlandır ---
    start_row = 11
    # toplam satırını bul
    total_row = None
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip().lower() == "toplam":
            total_row = r
            break
    if total_row is None:
        total_row = ws.max_row + 1

    current_capacity = total_row - start_row
    n_rows = len(rows)

    # Gerekirse satır ekle/sil
    if n_rows > current_capacity:
        add = n_rows - current_capacity
        ws.insert_rows(total_row, amount=add)
        # eklenen satırlara stil bas (DT/ODT satır alternansı)
        for i in range(add):
            dst_r = total_row + i
            src_r = start_row if ((dst_r - start_row) % 2 == 0) else (start_row + 1)
            _copy_row_style(ws, src_r, dst_r, 2, 40)  # B..AN
    elif n_rows < current_capacity:
        remove = current_capacity - n_rows
        ws.delete_rows(start_row + n_rows, amount=remove)

    # Yeni total satırı konumu
    total_row = start_row + n_rows
    last_data_row = total_row - 1

    # --- Satırları yaz ---
    col_channel = 2      # B
    col_group = 3        # C
    col_dtodt = 4        # D
    col_ratio = 5        # E
    col_month_adet = 37  # AK
    col_month_sn = 38    # AL
    col_unit_price = 39  # AM
    col_budget = 40      # AN

    for i, rr in enumerate(rows):
        r = start_row + i

        ws.cell(r, col_channel).value = str(rr.get("channel", "") or "")
        ws.cell(r, col_group).value = str(rr.get("publish_group", "") or "")
        ws.cell(r, col_dtodt).value = str(rr.get("dt_odt", "") or "")
        ws.cell(r, col_ratio).value = str(rr.get("dinlenme_orani", "NA") or "NA")

        # Günler
        day_vals = rr.get("days") or []
        for d in range(1, 32):
            c = day_start_col + (d - 1)
            if d <= len(day_vals):
                v = day_vals[d - 1]
                try:
                    ws.cell(r, c).value = int(v) if str(v).strip() != "" else None
                except Exception:
                    ws.cell(r, c).value = None
            else:
                ws.cell(r, c).value = None

        # Birim sn
        unit = rr.get("unit_price", 0) or 0
        try:
            ws.cell(r, col_unit_price).value = float(unit)
        except Exception:
            ws.cell(r, col_unit_price).value = 0.0

        # Formüller (Excel açılınca hesaplanacak)
        ws.cell(r, col_month_adet).value = f"=SUM(F{r}:AJ{r})"
        ws.cell(r, col_month_sn).value = f"=AK{r}*$C$8"
        ws.cell(r, col_budget).value = f"=AL{r}*AM{r}"

    # --- Toplam satırı ---
    ws.cell(total_row, col_channel).value = "Toplam"
    ws.cell(total_row, col_group).value = None
    ws.cell(total_row, col_dtodt).value = None
    ws.cell(total_row, col_ratio).value = None

    if last_data_row >= start_row:
        # Gün toplamları
        for d in range(1, 32):
            c = day_start_col + (d - 1)
            col_letter = get_column_letter(c)
            ws.cell(total_row, c).value = f"=SUM({col_letter}{start_row}:{col_letter}{last_data_row})"

        ws.cell(total_row, col_month_adet).value = f"=SUM(AK{start_row}:AK{last_data_row})"
        ws.cell(total_row, col_month_sn).value = f"=SUM(AL{start_row}:AL{last_data_row})"
        ws.cell(total_row, col_unit_price).value = None
        ws.cell(total_row, col_budget).value = f"=SUM(AN{start_row}:AN{last_data_row})"
    else:
        for d in range(1, 32):
            ws.cell(total_row, day_start_col + (d - 1)).value = None
        ws.cell(total_row, col_month_adet).value = 0
        ws.cell(total_row, col_month_sn).value = 0
        ws.cell(total_row, col_unit_price).value = None
        ws.cell(total_row, col_budget).value = 0

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def export_plan_ozet_range(out_path, data: dict) -> None:
    """PLAN ÖZET excel çıktısı (tarih aralığı bazlı, dinamik gün kolonları + ay kolonları)."""
    from copy import copy
    out_path = Path(out_path)

    template_path = resource_path("assets/plan_ozet_template.xlsx")
    if not template_path.exists():
        raise FileNotFoundError(f"Plan özet şablonu bulunamadı: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = wb["PLAN ÖZET"] if "PLAN ÖZET" in wb.sheetnames else wb.active

    header = data.get("header") or {}
    rows = data.get("rows") or []
    dates = data.get("dates") or []
    months = data.get("months") or []

    # --- Üst bilgiler ---
    ws["C2"].value = str(header.get("agency", "") or "")
    ws["C3"].value = str(header.get("advertiser", "") or "")
    ws["C4"].value = str(header.get("product", "") or "")
    ws["C5"].value = str(header.get("plan_title", "") or "")
    ws["C6"].value = str(header.get("reservation_no", "") or "")
    ws["C7"].value = str(header.get("period", "") or "")

    spot_len = header.get("spot_len", 0) or 0
    try:
        ws["C8"].value = float(spot_len)
    except Exception:
        ws["C8"].value = 0

    header_row = 10
    start_row = 11
    day_start_col = 6  # F

    def _find_col(value: str) -> int | None:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip() == value:
                return c
        return None

    # ay kolonlarının başladığı yer (ilk '* Adet' kolonu)
    month_start_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip().endswith("Adet"):
            month_start_col = c
            break

    unit_col = _find_col("Birim sn. (TL)")
    budget_col = _find_col("Toplam Bütçe\nNet TL") or _find_col("Toplam Bütçe")

    if month_start_col is None or unit_col is None or budget_col is None:
        raise RuntimeError("plan_ozet_template.xlsx başlıkları beklenen formatta değil.")

    def _copy_col_style(src_c: int, dst_c: int, min_r: int, max_r: int) -> None:
        for r in range(min_r, max_r + 1):
            s = ws.cell(r, src_c)
            d = ws.cell(r, dst_c)
            d._style = copy(s._style)
            d.number_format = s.number_format
            d.font = copy(s.font)
            d.border = copy(s.border)
            d.fill = copy(s.fill)
            d.alignment = copy(s.alignment)
            d.protection = copy(s.protection)
            d.comment = None

    def _copy_row_style(src_r: int, dst_r: int, min_c: int, max_c: int) -> None:
        ws.row_dimensions[dst_r].height = ws.row_dimensions[src_r].height
        for c in range(min_c, max_c + 1):
            s = ws.cell(src_r, c)
            d = ws.cell(dst_r, c)
            d._style = copy(s._style)
            d.number_format = s.number_format
            d.font = copy(s.font)
            d.border = copy(s.border)
            d.fill = copy(s.fill)
            d.alignment = copy(s.alignment)
            d.protection = copy(s.protection)
            d.comment = None

    # --- Gün kolonlarını aralığa göre ayarla ---
    need_days = int(len(dates))
    template_days = int(month_start_col - day_start_col)

    if need_days > template_days:
        add = need_days - template_days
        ws.insert_cols(month_start_col, amount=add)
        # eklenen kolonların stili: son gün kolonu
        src_c = month_start_col - 1
        for i in range(add):
            dst_c = month_start_col + i
            _copy_col_style(src_c, dst_c, header_row, ws.max_row)
    elif need_days < template_days:
        remove = template_days - need_days
        ws.delete_cols(month_start_col - remove, amount=remove)

    # kolonlar kaydı, yeniden bul
    month_start_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip().endswith("Adet"):
            month_start_col = c
            break
    unit_col = _find_col("Birim sn. (TL)")
    budget_col = _find_col("Toplam Bütçe\nNet TL") or _find_col("Toplam Bütçe")

    # --- Ay kolonlarını aralığa göre ayarla (her ay: Adet+Saniye) ---
    need_month_cols = int(len(months) * 2)
    template_month_cols = int(unit_col - month_start_col)

    if need_month_cols > template_month_cols:
        add = need_month_cols - template_month_cols
        ws.insert_cols(unit_col, amount=add)
        # stil: mevcut son ay kolonu
        src_c = unit_col - 1
        for i in range(add):
            dst_c = unit_col + i
            _copy_col_style(src_c, dst_c, header_row, ws.max_row)
    elif need_month_cols < template_month_cols:
        remove = template_month_cols - need_month_cols
        ws.delete_cols(unit_col - remove, amount=remove)

    # yeniden bul (unit/budget shift)
    month_start_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip().endswith("Adet"):
            month_start_col = c
            break
    unit_col = _find_col("Birim sn. (TL)")
    budget_col = _find_col("Toplam Bütçe\nNet TL") or _find_col("Toplam Bütçe")

    max_write_col = int(budget_col)

    # --- Başlıkları yaz ---
    # gün başlıkları: Pt\n09.03
    for i, dd in enumerate(dates):
        dow = TR_DOW[int(dd.weekday())]
        ws.cell(header_row, day_start_col + i).value = f"{dow}\n{dd:%d.%m}"
        ws.cell(header_row, day_start_col + i).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # fazla kolonlar varsa temizle (artık yok ama güvenlik)
    for c in range(day_start_col + need_days, month_start_col):
        ws.cell(header_row, c).value = None

    # ay başlıkları
    for i, (yy, mm) in enumerate(months):
        mname = ["OCAK","ŞUBAT","MART","NİSAN","MAYIS","HAZİRAN","TEMMUZ","AĞUSTOS","EYLÜL","EKİM","KASIM","ARALIK"][int(mm)-1]
        ws.cell(header_row, month_start_col + 2*i).value = f"{mname} Adet"
        ws.cell(header_row, month_start_col + 2*i + 1).value = f"{mname} Saniye"

    # --- Satır sayısını ayarla (Toplam satırını sabit tut) ---
    total_row = None
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip().lower() == "toplam":
            total_row = r
            break
    if total_row is None:
        total_row = ws.max_row + 1

    current_capacity = total_row - start_row
    n_rows = len(rows)

    if n_rows > current_capacity:
        add = n_rows - current_capacity
        ws.insert_rows(total_row, amount=add)
        for i in range(add):
            dst_r = total_row + i
            src_r = start_row if ((dst_r - start_row) % 2 == 0) else (start_row + 1)
            _copy_row_style(src_r, dst_r, 2, max_write_col)
    elif n_rows < current_capacity:
        remove = current_capacity - n_rows
        ws.delete_rows(start_row + n_rows, amount=remove)

    total_row = start_row + n_rows
    last_data_row = total_row - 1

    # --- Yaz ---
    col_channel = 2
    col_group = 3
    col_dtodt = 4
    col_ratio = 5

    for i, rr in enumerate(rows):
        r = start_row + i
        ws.cell(r, col_channel).value = str(rr.get("channel", "") or "")
        ws.cell(r, col_group).value = str(rr.get("publish_group", "") or "")
        ws.cell(r, col_dtodt).value = str(rr.get("dt_odt", "") or "")
        ws.cell(r, col_ratio).value = str(rr.get("dinlenme_orani", "NA") or "NA")

        dvals = rr.get("days") or []
        for j in range(need_days):
            v = dvals[j] if j < len(dvals) else ""
            if v in ("", None):
                ws.cell(r, day_start_col + j).value = None
            else:
                try:
                    ws.cell(r, day_start_col + j).value = int(v)
                except Exception:
                    ws.cell(r, day_start_col + j).value = None

        mcols = rr.get("month_cols") or []
        for j in range(need_month_cols):
            v = mcols[j] if j < len(mcols) else ""
            c = month_start_col + j
            if v in ("", None):
                ws.cell(r, c).value = None
            else:
                try:
                    ws.cell(r, c).value = int(v) if (j % 2 == 0) else float(v)
                except Exception:
                    ws.cell(r, c).value = v

        # unit price
        unit = rr.get("unit_price", "")
        if unit in ("", None):
            ws.cell(r, unit_col).value = None
        elif isinstance(unit, str):
            ws.cell(r, unit_col).value = unit
        else:
            try:
                ws.cell(r, unit_col).value = float(unit)
            except Exception:
                ws.cell(r, unit_col).value = unit

        bud = rr.get("budget", "")
        if bud in ("", None):
            ws.cell(r, budget_col).value = None
        else:
            try:
                ws.cell(r, budget_col).value = float(bud)
            except Exception:
                ws.cell(r, budget_col).value = bud

    # --- Toplam satırı ---
    ws.cell(total_row, col_channel).value = "Toplam"
    ws.cell(total_row, col_group).value = None
    ws.cell(total_row, col_dtodt).value = None
    ws.cell(total_row, col_ratio).value = None

    totals = data.get("totals") or {}
    tdays = totals.get("days") or []
    for j in range(need_days):
        v = tdays[j] if j < len(tdays) else ""
        ws.cell(total_row, day_start_col + j).value = (None if v in ("", None) else int(v))

    tm = totals.get("month_cols") or []
    for j in range(need_month_cols):
        v = tm[j] if j < len(tm) else ""
        c = month_start_col + j
        if v in ("", None):
            ws.cell(total_row, c).value = None
        else:
            ws.cell(total_row, c).value = int(v) if (j % 2 == 0) else float(v)

    ws.cell(total_row, unit_col).value = None
    tb = totals.get("budget", "")
    ws.cell(total_row, budget_col).value = None if tb in ("", None) else float(tb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
